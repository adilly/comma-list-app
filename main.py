from openpyxl import load_workbook, workbook

# Execution python -c 'import main; main.convertsql(1)'

def convertsql(dataType):
    #dataType from manual? 
    # Can you pull dataType based on column value? 
    wb = load_workbook("D:\Git\sql-convert-file.xlsx")
    # parm file name from clip baord
    # source from xlsx or txt or made copy command???
    ws = wb.active
    firstColumn = ws['A']

    if type(dataType) is str:
        for x in range(len(firstColumn)):
            print(firstColumn[x].value + ' ')
    elif type(dataType) is int:
        for x in range(len(firstColumn)):
            print(firstColumn[x].value)
    else:
        print('This is not a valid SQL type.')
# else dataType = junk
    # then junk

    

